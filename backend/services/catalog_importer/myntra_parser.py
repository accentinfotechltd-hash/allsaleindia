"""Myntra catalog export parser.

Myntra sellers download their listing from Myntra Partner Portal as an
Excel (.xlsx) with columns:
    Style Code | Vendor Article Number | Article Type | Brand |
    Color | Size | MRP | Selling Price | Inventory |
    Title | Description | Composition | Wash Care | HSN |
    Front Image, Back Image, Side Image, Detail Image, Look Image, ...

We map those into ``MappedProduct`` exactly the same way Amazon/Flipkart
parsers do, so the existing /preview + /commit flow keeps working with
zero downstream changes.

Detection: handled in ``detect.py``. We recognise either filename hint
("myntra") or the column-header signature {"Style Code", "Article Type",
"Selling Price"} appearing in the first sheet.
"""
from __future__ import annotations

import io
from typing import Any, Dict, List, Optional

import openpyxl  # type: ignore

from .mapping import coerce_inr_to_nzd, parse_decimal, parse_int, split_multi
from .models import MappedProduct, ParsedCatalog, ParsedRow, RowIssue


# Myntra `Article Type` → Allsale (category, subcategory).
_MYNTRA_ARTICLE_TO_ALLSALE: Dict[str, tuple[str, Optional[str]]] = {
    "kurta": ("Ethnic Fashion", "Kurtis"),
    "kurta sets": ("Ethnic Fashion", "Kurta Sets"),
    "kurti": ("Ethnic Fashion", "Kurtis"),
    "sarees": ("Ethnic Fashion", "Sarees"),
    "saree": ("Ethnic Fashion", "Sarees"),
    "lehenga choli": ("Ethnic Fashion", "Lehengas"),
    "tshirts": ("Men's Clothing", "T-Shirts"),
    "t-shirts": ("Men's Clothing", "T-Shirts"),
    "shirts": ("Men's Clothing", "Shirts"),
    "jeans": ("Men's Clothing", "Bottoms"),
    "dresses": ("Women's Clothing", "Dresses"),
    "tops": ("Women's Clothing", "Tops"),
    "heels": ("Shoes", "Heels"),
    "flats": ("Shoes", "Flats"),
    "sneakers": ("Shoes", "Sneakers"),
    "earrings": ("Jewellery", "Earrings"),
    "necklaces": ("Jewellery", "Necklaces"),
    "bangles": ("Jewellery", "Bangles"),
    "watches": ("Accessories", "Watches"),
    "wallets": ("Accessories", "Wallets"),
    "handbags": ("Accessories", "Handbags"),
}


def _map_article_type(at: str | None) -> tuple[Optional[str], Optional[str]]:
    if not at:
        return None, None
    k = at.strip().lower()
    if k in _MYNTRA_ARTICLE_TO_ALLSALE:
        return _MYNTRA_ARTICLE_TO_ALLSALE[k]
    for hit_key, val in _MYNTRA_ARTICLE_TO_ALLSALE.items():
        if hit_key in k or k in hit_key:
            return val
    # Coarse fallbacks
    if "shirt" in k:
        return "Men's Clothing", "Tops"
    if "shoe" in k:
        return "Shoes", None
    return None, None


def _norm_headers(row: List[Any]) -> List[str]:
    return [(str(c) if c is not None else "").strip().lower() for c in row]


def _idx(headers: List[str], *candidates: str) -> Optional[int]:
    """Find a column index whose normalised header matches any candidate."""
    for c in candidates:
        c = c.strip().lower()
        if c in headers:
            return headers.index(c)
    # Substring fallback for fuzzy match
    for c in candidates:
        c = c.strip().lower()
        for i, h in enumerate(headers):
            if c in h:
                return i
    return None


def parse_myntra(
    file_bytes: bytes, *, fx_inr_per_nzd: float = 51.0
) -> ParsedCatalog:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)

    # Headers — Myntra sometimes has a metadata row before the headers; we
    # scan up to the first 5 rows to find the one containing "style code".
    headers: List[str] = []
    for _ in range(5):
        try:
            candidate = next(rows_iter)
        except StopIteration:
            break
        norm = _norm_headers(list(candidate))
        if any("style code" in h or "style_code" in h for h in norm):
            headers = norm
            break

    if not headers:
        raise ValueError("Couldn't find Myntra header row (expected 'Style Code')")

    col_style = _idx(headers, "style code", "style id")
    col_vendor_sku = _idx(headers, "vendor article number", "vendor sku", "sku")
    col_brand = _idx(headers, "brand")
    col_article = _idx(headers, "article type", "category")
    col_color = _idx(headers, "color", "colour")
    col_size = _idx(headers, "size")
    col_mrp = _idx(headers, "mrp")
    col_price = _idx(headers, "selling price", "net selling price", "vendor price")
    col_stock = _idx(headers, "inventory", "stock", "qty")
    col_title = _idx(headers, "title", "product name", "name")
    col_desc = _idx(headers, "description")
    col_composition = _idx(headers, "composition", "fabric")
    col_wash = _idx(headers, "wash care")
    col_hsn = _idx(headers, "hsn", "hsn code")
    # Myntra puts up to ~6 image columns
    image_cols: List[int] = [
        i
        for i, h in enumerate(headers)
        if "image" in h or "look" in h or "front" in h or "back" in h
    ]

    if col_title is None and col_style is None:
        raise ValueError("Myntra sheet missing Title and Style Code columns")

    parsed_rows: List[ParsedRow] = []
    for r_idx, row in enumerate(rows_iter, start=len(parsed_rows) + 2):
        row_l = list(row)
        # Skip fully empty rows
        if not any(v is not None and str(v).strip() for v in row_l):
            continue

        def cell(i: Optional[int]) -> Any:
            return row_l[i] if i is not None and i < len(row_l) else None

        title = str(cell(col_title) or cell(col_style) or "").strip()
        if not title:
            continue  # Don't add a row that's basically empty

        issues: List[RowIssue] = []
        price = parse_decimal(cell(col_price))
        mrp = parse_decimal(cell(col_mrp))
        if price is None and mrp is not None:
            price = mrp
        if price is None:
            issues.append(
                RowIssue(severity="error", field="price_inr", message="Selling Price required")
            )

        # Images
        gallery: List[str] = []
        for ic in image_cols:
            v = cell(ic)
            if v and str(v).strip().startswith("http"):
                gallery.append(str(v).strip())
        if not gallery:
            issues.append(
                RowIssue(severity="error", field="image", message="No image URLs in row")
            )

        cat, sub = _map_article_type(str(cell(col_article) or ""))
        if not cat:
            issues.append(
                RowIssue(
                    severity="warning",
                    field="category",
                    message="Article type didn't map to an Allsale category — pick manually",
                )
            )

        # Build description: join composition + wash care + description fields.
        desc_parts: List[str] = []
        if cell(col_desc):
            desc_parts.append(str(cell(col_desc)).strip())
        if cell(col_composition):
            desc_parts.append(f"Fabric: {str(cell(col_composition)).strip()}")
        if cell(col_wash):
            desc_parts.append(f"Wash care: {str(cell(col_wash)).strip()}")
        description = " · ".join([p for p in desc_parts if p])

        prod = MappedProduct(
            sku=str(cell(col_vendor_sku) or cell(col_style) or "").strip() or None,
            name=title,
            description=description,
            category=cat or "",
            subcategory=sub,
            brand=str(cell(col_brand) or "").strip() or None,
            price_inr=price,
            price_nzd=coerce_inr_to_nzd(price, fx_inr_per_nzd) if price else None,
            mrp_inr=mrp,
            stock_count=parse_int(cell(col_stock)) or 0,
            image=gallery[0] if gallery else None,
            images=gallery[1:] if len(gallery) > 1 else [],
            bullets=[],
            colors=split_multi(cell(col_color), seps=(",", "/", ";")),
            sizes=split_multi(cell(col_size), seps=(",", "/", ";", "|")),
            weight_kg=None,
            hsn_code=str(cell(col_hsn) or "").strip() or None,
            ean_upc=None,
            country_of_origin="India",
            manufacturer=None,
            ingredients=None,
            raw_category_label=str(cell(col_article) or "").strip() or None,
        )
        ready = not any(i.severity == "error" for i in issues)
        parsed_rows.append(
            ParsedRow(row_index=r_idx, product=prod, issues=issues, ready=ready)
        )

    ready_count = sum(1 for r in parsed_rows if r.ready)
    needs = len(parsed_rows) - ready_count
    return ParsedCatalog(
        source="myntra",
        total_rows=len(parsed_rows),
        ready_count=ready_count,
        needs_attention_count=needs,
        rows=parsed_rows,
        sheet_name=ws.title,
        fx_inr_to_nzd=fx_inr_per_nzd,
    )
