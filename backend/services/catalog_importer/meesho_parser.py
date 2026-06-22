"""Meesho catalog export parser.

Meesho's seller panel offers two export shapes:
  • "Bulk Listing Template" — Excel with a fixed-column schema
  • "My Inventory.csv" — comma-separated, slightly different column names

We support both via the same parser; detection in ``detect.py`` looks for
either filename hints ("meesho") OR header signatures like
"Product ID" + "Price" + "Inventory".

Expected columns (case-insensitive, ANY of these aliases are accepted):
  Product ID | SKU
  Title | Product Name
  Category | Sub Category
  Description | Product Description
  MRP | Maximum Retail Price
  Price | Selling Price | Final Price
  Inventory | Stock | Quantity
  Image URL 1..7 | Photo 1..7
  Color | Size | Brand | HSN
"""
from __future__ import annotations

import io
from typing import Any, Dict, List, Optional

from .mapping import coerce_inr_to_nzd, parse_decimal, parse_int, split_multi
from .models import MappedProduct, ParsedCatalog, ParsedRow, RowIssue


_MEESHO_CATEGORY_TO_ALLSALE: Dict[str, tuple[str, Optional[str]]] = {
    "saree": ("Ethnic Fashion", "Sarees"),
    "sarees": ("Ethnic Fashion", "Sarees"),
    "kurta": ("Ethnic Fashion", "Kurtis"),
    "kurti": ("Ethnic Fashion", "Kurtis"),
    "lehenga": ("Ethnic Fashion", "Lehengas"),
    "suit": ("Ethnic Fashion", "Suits"),
    "men ethnic": ("Ethnic Fashion", None),
    "women ethnic": ("Ethnic Fashion", None),
    "western dresses": ("Women's Clothing", "Dresses"),
    "tops": ("Women's Clothing", "Tops"),
    "jewellery": ("Jewellery", None),
    "jewelry": ("Jewellery", None),
    "handbags": ("Accessories", "Handbags"),
    "watches": ("Accessories", "Watches"),
    "home & kitchen": ("Home & Kitchen", None),
    "kitchen": ("Home & Kitchen", "Kitchenware"),
    "beauty": ("Beauty & Health", None),
    "personal care": ("Beauty & Health", None),
    "mobiles": ("Electronics", None),
    "electronics": ("Electronics", None),
}


def _map_meesho_category(c: str | None) -> tuple[Optional[str], Optional[str]]:
    if not c:
        return None, None
    k = c.strip().lower()
    if k in _MEESHO_CATEGORY_TO_ALLSALE:
        return _MEESHO_CATEGORY_TO_ALLSALE[k]
    for hit_key, val in _MEESHO_CATEGORY_TO_ALLSALE.items():
        if hit_key in k or k in hit_key:
            return val
    return None, None


def _idx(headers: List[str], *candidates: str) -> Optional[int]:
    for c in candidates:
        c = c.strip().lower()
        if c in headers:
            return headers.index(c)
    for c in candidates:
        c = c.strip().lower()
        for i, h in enumerate(headers):
            if c in h:
                return i
    return None


def _read_rows(file_bytes: bytes) -> tuple[List[str], List[List[Any]], str]:
    """Read either an .xlsx or a .csv and return (headers, data_rows, sheet_name)."""
    head = file_bytes[:8]
    is_xlsx = head.startswith(b"PK\x03\x04")
    if is_xlsx:
        import openpyxl  # type: ignore

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], [], ws.title
        headers = [(str(c) if c is not None else "").strip().lower() for c in rows[0]]
        return headers, [list(r) for r in rows[1:]], ws.title

    # CSV
    import csv

    text = file_bytes.decode("utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:2048], delimiters=",;\t|")
    except Exception:  # noqa: BLE001
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect=dialect)
    rows = list(reader)
    if not rows:
        return [], [], "csv"
    headers = [(c or "").strip().lower() for c in rows[0]]
    return headers, rows[1:], "csv"


def parse_meesho(
    file_bytes: bytes, *, fx_inr_per_nzd: float = 51.0
) -> ParsedCatalog:
    headers, data_rows, sheet_name = _read_rows(file_bytes)
    if not headers:
        raise ValueError("Empty Meesho file")

    col_sku = _idx(headers, "product id", "sku", "product code")
    col_title = _idx(headers, "title", "product name", "name")
    col_desc = _idx(headers, "description", "product description")
    col_category = _idx(headers, "category", "main category")
    col_sub = _idx(headers, "sub category", "sub-category", "subcategory")
    col_mrp = _idx(headers, "mrp", "maximum retail price")
    col_price = _idx(headers, "selling price", "final price", "price")
    col_stock = _idx(headers, "inventory", "stock", "quantity", "qty")
    col_brand = _idx(headers, "brand")
    col_color = _idx(headers, "color", "colour")
    col_size = _idx(headers, "size")
    col_hsn = _idx(headers, "hsn", "hsn code")
    # Image columns — Meesho uses Image URL 1..7 OR Photo 1..7
    image_cols = [
        i for i, h in enumerate(headers)
        if ("image url" in h or h.startswith("photo") or "main image" in h or "product image" in h)
    ]

    if col_title is None and col_sku is None:
        raise ValueError("Meesho file missing Title and Product ID columns")

    parsed_rows: List[ParsedRow] = []
    for r_idx, row_l in enumerate(data_rows, start=2):
        if not any(v not in (None, "") for v in row_l):
            continue

        def cell(i: Optional[int]) -> Any:
            return row_l[i] if i is not None and i < len(row_l) else None

        title = str(cell(col_title) or cell(col_sku) or "").strip()
        if not title:
            continue
        issues: List[RowIssue] = []
        price = parse_decimal(cell(col_price))
        mrp = parse_decimal(cell(col_mrp))
        if price is None and mrp is not None:
            price = mrp
        if price is None:
            issues.append(
                RowIssue(severity="error", field="price_inr", message="Selling Price required")
            )

        gallery: List[str] = []
        for ic in image_cols:
            v = cell(ic)
            if v and str(v).strip().startswith("http"):
                gallery.append(str(v).strip())
        if not gallery:
            issues.append(
                RowIssue(severity="error", field="image", message="No image URLs in row")
            )

        cat, sub = _map_meesho_category(str(cell(col_category) or ""))
        if not cat:
            issues.append(
                RowIssue(
                    severity="warning",
                    field="category",
                    message="Meesho category didn't map to Allsale — review",
                )
            )
        if not sub and cell(col_sub):
            sub = str(cell(col_sub)).strip() or None

        prod = MappedProduct(
            sku=str(cell(col_sku) or "").strip() or None,
            name=title,
            description=str(cell(col_desc) or "").strip(),
            category=cat or "",
            subcategory=sub,
            brand=str(cell(col_brand) or "").strip() or None,
            price_inr=price,
            price_nzd=coerce_inr_to_nzd(price, fx_inr_per_nzd) if price else None,
            mrp_inr=mrp,
            stock_count=parse_int(cell(col_stock)) or 0,
            image=gallery[0] if gallery else None,
            images=gallery[1:] if len(gallery) > 1 else [],
            bullets=[],
            colors=split_multi(cell(col_color), seps=(",", "/", ";")),
            sizes=split_multi(cell(col_size), seps=(",", "/", ";", "|")),
            weight_kg=None,
            hsn_code=str(cell(col_hsn) or "").strip() or None,
            ean_upc=None,
            country_of_origin="India",
            manufacturer=None,
            ingredients=None,
            raw_category_label=str(cell(col_category) or "").strip() or None,
        )
        ready = not any(i.severity == "error" for i in issues)
        parsed_rows.append(
            ParsedRow(row_index=r_idx, product=prod, issues=issues, ready=ready)
        )

    ready_count = sum(1 for r in parsed_rows if r.ready)
    needs = len(parsed_rows) - ready_count
    return ParsedCatalog(
        source="meesho",
        total_rows=len(parsed_rows),
        ready_count=ready_count,
        needs_attention_count=needs,
        rows=parsed_rows,
        sheet_name=sheet_name,
        fx_inr_to_nzd=fx_inr_per_nzd,
    )
