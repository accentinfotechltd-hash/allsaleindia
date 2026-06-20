"""Parse Flipkart catalog .xls (legacy OLE) / .xlsx.

File structure (validated against the real seller-uploaded template):
  Each category has its own sheet named like ``conditioner``, ``saree``
  etc. Within that sheet:
    - Row 0: column header labels (e.g. "Seller SKU ID", "MRP (INR)")
    - Row 1: type hints ("Single - Text", "URL", "Single - Decimal")
    - Row 2: example row
    - Row 3: description (skip)
    - Row 4+: actual seller-entered product data

We identify the "data" sheet by skipping known meta sheets and picking
the first sheet whose ncols > 30 and that contains a "Seller SKU ID" or
"Product Title" header.
"""
from __future__ import annotations

import io
import logging

from .mapping import (
    coerce_inr_to_nzd,
    map_flipkart_sheet,
    parse_decimal,
    parse_int,
    split_multi,
)
from .models import MappedProduct, ParsedCatalog, ParsedRow, RowIssue

log = logging.getLogger("allsale.importer.flipkart")

META_SHEETS = {
    "summary sheet",
    "index",
    "listing faq sheet",
    "image guidelines",
    "matchingattributes",
    "variantattributes",
    "parent variant products",
    "template_version",
}


def _open_book(file_bytes: bytes):
    """Try .xls first (legacy), then .xlsx as fallback."""
    head = file_bytes[:8]
    if head.startswith(b"\xD0\xCF\x11\xE0"):
        import xlrd

        wb = xlrd.open_workbook(file_contents=file_bytes, on_demand=False)

        class XlrdAdapter:
            def __init__(self, wb):
                self._wb = wb

            @property
            def sheetnames(self):
                return self._wb.sheet_names()

            def sheet(self, name):
                return XlrdSheetAdapter(self._wb.sheet_by_name(name))

        class XlrdSheetAdapter:
            def __init__(self, s):
                self._s = s

            @property
            def nrows(self):
                return self._s.nrows

            @property
            def ncols(self):
                return self._s.ncols

            def row(self, ri: int):
                if ri >= self._s.nrows:
                    return []
                return [self._s.cell_value(ri, ci) for ci in range(self._s.ncols)]

        return XlrdAdapter(wb)

    # .xlsx fallback
    import openpyxl

    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes), data_only=True, read_only=True
    )

    class OpAdapter:
        def __init__(self, wb):
            self._wb = wb

        @property
        def sheetnames(self):
            return self._wb.sheetnames

        def sheet(self, name):
            return OpSheetAdapter(self._wb[name])

    class OpSheetAdapter:
        def __init__(self, s):
            self._s = s
            self._rows = list(s.iter_rows(values_only=True))

        @property
        def nrows(self):
            return len(self._rows)

        @property
        def ncols(self):
            return max((len(r) for r in self._rows), default=0)

        def row(self, ri: int):
            if ri >= len(self._rows):
                return []
            return list(self._rows[ri])

    return OpAdapter(wb)


def _find_data_sheet(wb) -> str:
    candidates = []
    for name in wb.sheetnames:
        if name.strip().lower() in META_SHEETS:
            continue
        if name.lower().startswith("dropdownvaluesfor"):
            continue
        s = wb.sheet(name)
        if s.ncols < 20 or s.nrows < 2:
            continue
        # The data sheet's row-0 should contain canonical Flipkart fields.
        row0 = [str(v).strip().lower() for v in s.row(0)]
        joined = " ".join(row0)
        if (
            "seller sku id" in joined
            or "mrp" in joined
            or "main image url" in joined
        ):
            candidates.append((name, s.ncols))
    if not candidates:
        raise ValueError(
            "Could not find a Flipkart product data sheet. "
            "Expected a sheet like 'conditioner' / 'saree' with seller SKU / MRP columns."
        )
    # Prefer the widest (most attributes).
    candidates.sort(key=lambda x: -x[1])
    return candidates[0][0]


def parse_flipkart(file_bytes: bytes, *, fx_inr_per_nzd: float = 51.0) -> ParsedCatalog:
    wb = _open_book(file_bytes)
    sheet_name = _find_data_sheet(wb)
    s = wb.sheet(sheet_name)

    headers = [str(v).strip() if v is not None else "" for v in s.row(0)]

    def col_idx(label: str) -> int | None:
        for i, h in enumerate(headers):
            if h.lower() == label.lower():
                return i
        return None

    parsed_rows: list[ParsedRow] = []
    # Data starts at row 4 (skip header, type-hint, example, description).
    for ri in range(4, s.nrows):
        row = s.row(ri)
        if not any(v not in (None, "") for v in row):
            continue

        def get(label: str):
            i = col_idx(label)
            if i is None or i >= len(row):
                return None
            v = row[i]
            if v in (None, ""):
                return None
            return v

        sku = str(get("Seller SKU ID") or "").strip() or None
        brand = str(get("Brand") or "").strip() or None
        model = str(get("Model Name") or "").strip() or None
        # Flipkart split: title = Brand + Model. Provide best-effort fallback.
        name = " ".join([p for p in (brand, model) if p]).strip()

        issues: list[RowIssue] = []
        if not name:
            issues.append(
                RowIssue(severity="error", field="name", message="Brand + Model Name both empty")
            )

        main_img = str(get("Main Image URL") or "").strip() or None
        gallery: list[str] = []
        for label in (
            "Other Image URL 1",
            "Other Image URL 2",
            "Other Image URL 3",
            "Other Image URL 4",
        ):
            u = get(label)
            if u:
                gallery.append(str(u).strip())
        if not main_img and not gallery:
            issues.append(
                RowIssue(severity="error", field="image", message="No image URL found")
            )

        price_inr = parse_decimal(get("Your selling price (INR)"))
        mrp_inr = parse_decimal(get("MRP (INR)"))
        if price_inr is None:
            issues.append(
                RowIssue(
                    severity="error",
                    field="price_inr",
                    message="Your selling price (INR) is empty",
                )
            )
        price_nzd = (
            coerce_inr_to_nzd(price_inr, fx_inr_per_nzd)
            if price_inr is not None
            else None
        )

        stock = parse_int(get("Stock")) or 0
        description = str(get("Description") or "").strip()
        bullets = split_multi(get("Key Features"), seps=("::", "|", ";"))
        if not description and not bullets:
            issues.append(
                RowIssue(
                    severity="warning",
                    field="description",
                    message="No description or key features",
                )
            )

        # Category from sheet name
        cat, subcat = map_flipkart_sheet(sheet_name)
        if cat is None:
            issues.append(
                RowIssue(
                    severity="warning",
                    field="category",
                    message=(
                        f"Unmapped sheet '{sheet_name}'. Pick a category before publishing."
                    ),
                )
            )

        # Dimensions
        L = parse_decimal(get("Length (CM)"))
        B = parse_decimal(get("Breadth (CM)"))
        H = parse_decimal(get("Height (CM)"))
        dims = [L or 0.0, B or 0.0, H or 0.0] if any((L, B, H)) else None
        weight = parse_decimal(get("Weight (KG)"))

        prod = MappedProduct(
            sku=sku,
            name=name or "(unnamed)",
            description=description,
            category=cat or "",
            subcategory=subcat,
            brand=brand,
            price_inr=price_inr,
            price_nzd=price_nzd,
            mrp_inr=mrp_inr,
            stock_count=stock,
            image=main_img or (gallery[0] if gallery else None),
            images=gallery if main_img else gallery[1:],
            bullets=bullets,
            weight_kg=weight,
            dimensions_cm=dims,
            hsn_code=str(get("HSN") or "").strip() or None,
            ean_upc=str(get("EAN/UPC") or "").strip() or None,
            country_of_origin=str(get("Country Of Origin") or "India").strip(),
            manufacturer=str(get("Manufacturer Details") or "").strip() or None,
            importer=str(get("Importer Details") or "").strip() or None,
            ingredients=str(get("Ingredients") or "").strip() or None,
            shelf_life_months=parse_int(get("Shelf Life (MONTHS)")),
            raw_category_label=sheet_name,
        )

        ready = not any(i.severity == "error" for i in issues)
        parsed_rows.append(
            ParsedRow(row_index=ri + 1, product=prod, issues=issues, ready=ready)
        )

    ready_count = sum(1 for r in parsed_rows if r.ready)
    needs = len(parsed_rows) - ready_count

    return ParsedCatalog(
        source="flipkart",
        total_rows=len(parsed_rows),
        ready_count=ready_count,
        needs_attention_count=needs,
        rows=parsed_rows,
        sheet_name=sheet_name,
        fx_inr_to_nzd=fx_inr_per_nzd,
        warnings=[],
    )
