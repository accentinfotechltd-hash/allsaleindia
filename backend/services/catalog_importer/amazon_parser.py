"""Parse Amazon seller catalog template (.xlsx / .xlsm).

File structure (validated against the real seller-uploaded template):
  Sheet "Template":
    - Row 4: human-readable column header (e.g. "Item Name", "Brand Name")
    - Row 5: internal field code (e.g. ``item_name[marketplace_id=...]``)
    - Row 6+: actual product rows

We use Row 4 as our mapping key (stable across marketplaces). The
column count varies per product type, so we extract by header label,
not by index.
"""
from __future__ import annotations

import io
import logging
from typing import Optional

import openpyxl

from .mapping import (
    coerce_inr_to_nzd,
    map_amazon_product_type,
    parse_decimal,
    parse_int,
    split_multi,
)
from .models import MappedProduct, ParsedCatalog, ParsedRow, RowIssue

log = logging.getLogger("allsale.importer.amazon")


def _row_to_dict(
    headers: list[str], row_vals: tuple
) -> dict:
    """Convert a sheet row to ``{header_label: value}``. Same header may
    repeat (e.g. "Bullet Point" ×5) — we collect repeats into a list.
    """
    out: dict = {}
    for h, v in zip(headers, row_vals):
        if h is None or h == "":
            continue
        if v is None or v == "":
            continue
        if h in out:
            # repeated header → collect into list
            if isinstance(out[h], list):
                out[h].append(v)
            else:
                out[h] = [out[h], v]
        else:
            out[h] = v
    return out


def parse_amazon(file_bytes: bytes, *, fx_inr_per_nzd: float = 51.0) -> ParsedCatalog:
    """Parse an Amazon listings .xlsm/.xlsx into a ``ParsedCatalog``."""
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), data_only=True, read_only=True, keep_vba=False
        )
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"Could not open Amazon workbook: {exc}") from exc

    if "Template" not in wb.sheetnames:
        raise ValueError(
            "This doesn't look like an Amazon catalog template (no 'Template' sheet)."
        )

    sheet = wb["Template"]
    rows_iter = sheet.iter_rows(values_only=True)

    # Skip rows 1, 2, 3 (control headers), read row 4 (human labels).
    for _ in range(3):
        next(rows_iter, None)
    header_row = next(rows_iter, None) or ()
    headers: list[str] = [
        (str(v).strip() if v is not None else "") for v in header_row
    ]
    # Skip row 5 (internal field codes).
    next(rows_iter, None)

    parsed_rows: list[ParsedRow] = []
    for idx, row in enumerate(rows_iter, start=6):
        if not any(v not in (None, "") for v in row):
            continue  # skip blank rows

        data = _row_to_dict(headers, row)

        # ------- Hard-required fields
        sku = str(data.get("SKU") or "").strip() or None
        name = str(data.get("Item Name") or "").strip()
        brand = str(data.get("Brand Name") or "").strip() or None
        product_type = str(data.get("Product Type") or "").strip() or None

        issues: list[RowIssue] = []
        if not name:
            issues.append(
                RowIssue(severity="error", field="name", message="Item Name is empty")
            )

        # ------- Images
        main_img = str(data.get("Main Image URL") or "").strip() or None
        other_imgs_raw = data.get("Other Image URL") or []
        if not isinstance(other_imgs_raw, list):
            other_imgs_raw = [other_imgs_raw]
        gallery = [str(u).strip() for u in other_imgs_raw if u]
        if not main_img and not gallery:
            issues.append(
                RowIssue(severity="error", field="image", message="No image URL found")
            )

        # ------- Description + bullets
        description = str(data.get("Product Description") or "").strip()
        bullets_raw = data.get("Bullet Point") or []
        if not isinstance(bullets_raw, list):
            bullets_raw = [bullets_raw]
        bullets = [str(b).strip() for b in bullets_raw if b]
        if not description and not bullets:
            issues.append(
                RowIssue(
                    severity="warning",
                    field="description",
                    message="No description or bullet points",
                )
            )

        # ------- Category mapping
        cat, subcat = map_amazon_product_type(product_type)
        if cat is None:
            issues.append(
                RowIssue(
                    severity="warning",
                    field="category",
                    message=(
                        f"Unmapped product_type '{product_type}'. Pick a category before publishing."
                    ),
                )
            )

        # ------- Price (Amazon templates often don't include price on the
        # listings sheet — it's on a separate Pricing sheet). Mark warning.
        # If a "Standard Price" or "List Price" column exists, use it.
        price_inr = (
            parse_decimal(data.get("Standard Price"))
            or parse_decimal(data.get("List Price"))
        )
        if price_inr is None:
            issues.append(
                RowIssue(
                    severity="warning",
                    field="price_inr",
                    message="No price in Amazon template. Set price before publishing.",
                )
            )
        price_nzd = (
            coerce_inr_to_nzd(price_inr, fx_inr_per_nzd)
            if price_inr is not None
            else None
        )

        # ------- Variants (color/size)
        colors = []
        if data.get("Color"):
            colors = split_multi(data.get("Color"), seps=(",", "|", "::"))
        if data.get("Color Map") and not colors:
            colors = split_multi(data.get("Color Map"), seps=(",", "|"))
        sizes = split_multi(data.get("Size"), seps=(",", "|")) if data.get("Size") else []

        # ------- Other attributes (best-effort)
        prod = MappedProduct(
            sku=sku,
            name=name or "(unnamed)",
            description=description,
            category=cat or "",
            subcategory=subcat,
            brand=brand,
            price_inr=price_inr,
            price_nzd=price_nzd,
            stock_count=parse_int(data.get("Quantity")) or 0,
            image=main_img or (gallery[0] if gallery else None),
            images=gallery if main_img else gallery[1:],
            bullets=bullets,
            colors=colors,
            sizes=sizes,
            hsn_code=str(
                data.get("External Product Information")
                or data.get("HSN Code")
                or ""
            ).strip()
            or None,
            ean_upc=str(data.get("Product Id") or "").strip() or None,
            country_of_origin=str(data.get("Country of Origin") or "India").strip(),
            manufacturer=str(data.get("Manufacturer") or "").strip() or None,
            importer=str(
                data.get("Importer Contact Information") or ""
            ).strip()
            or None,
            ingredients=" ".join(
                str(x) for x in (data.get("Ingredients") or []) if x
            )
            if isinstance(data.get("Ingredients"), list)
            else (str(data.get("Ingredients") or "").strip() or None),
            raw_category_label=product_type,
        )

        ready = not any(i.severity == "error" for i in issues)
        parsed_rows.append(
            ParsedRow(row_index=idx, product=prod, issues=issues, ready=ready)
        )

    ready_count = sum(1 for r in parsed_rows if r.ready)
    needs = len(parsed_rows) - ready_count

    return ParsedCatalog(
        source="amazon",
        total_rows=len(parsed_rows),
        ready_count=ready_count,
        needs_attention_count=needs,
        rows=parsed_rows,
        sheet_name="Template",
        fx_inr_to_nzd=fx_inr_per_nzd,
        warnings=[
            (
                "Amazon listing templates often ship pricing in a separate sheet. "
                "Set price + stock on each row before publishing if not detected."
            )
        ]
        if any(r.product.price_inr is None for r in parsed_rows)
        else [],
    )
