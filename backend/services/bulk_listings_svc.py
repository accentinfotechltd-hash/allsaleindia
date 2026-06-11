"""Bulk CSV/XLSX product listing import & export for sellers.

Handles parsing uploaded files into normalized listing rows, validating
each row, and producing a downloadable template / export file.
"""
from __future__ import annotations

import csv
import io
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook

# Canonical column names users see in their CSV/XLSX template.
# `product_id` is OPTIONAL — when set, the row UPDATES an existing
# listing (must belong to the seller); otherwise a new listing is
# inserted.
TEMPLATE_COLUMNS: list[str] = [
    "product_id",
    "name",
    "description",
    "category",
    "subcategory",
    "price_nzd",
    "stock_count",
    "sizes",
    "colors",
    "shipping_days_min",
    "shipping_days_max",
    "image_urls",
]

REQUIRED_FOR_NEW: list[str] = [
    "name",
    "description",
    "category",
    "price_nzd",
    "image_urls",
]

# Tolerate small whitespace differences in column headers from sellers
# editing the template in Excel/Numbers/Google Sheets.
HEADER_ALIASES: dict[str, str] = {
    "id": "product_id",
    "product id": "product_id",
    "title": "name",
    "product name": "name",
    "desc": "description",
    "price (nzd)": "price_nzd",
    "price_nz": "price_nzd",
    "price": "price_nzd",
    "stock": "stock_count",
    "qty": "stock_count",
    "quantity": "stock_count",
    "images": "image_urls",
    "image": "image_urls",
    "size": "sizes",
    "color": "colors",
}


def _normalize_header(h: Any) -> str:
    s = ("" if h is None else str(h)).strip().lower().replace("-", "_")
    s = "_".join(s.split())  # collapse whitespace
    # try alias first by display form
    display = s.replace("_", " ")
    if display in HEADER_ALIASES:
        return HEADER_ALIASES[display]
    if s in HEADER_ALIASES:
        return HEADER_ALIASES[s]
    return s


def _split_multi(value: Any, sep: str = "|") -> list[str]:
    """Split a multi-value field. Accepts `|` (preferred), `,` or `;`."""
    if value is None:
        return []
    s = str(value).strip()
    if not s:
        return []
    # If user used commas/semicolons, treat those as separators too.
    for ch in (";", ","):
        s = s.replace(ch, sep)
    return [t.strip() for t in s.split(sep) if t and t.strip()]


def _coerce_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _coerce_int(value: Any, default: int | None = None) -> int | None:
    if value is None or value == "":
        return default
    try:
        return int(float(str(value).replace(",", "").strip()))
    except (TypeError, ValueError):
        return default


def parse_csv_bytes(blob: bytes) -> list[dict[str, Any]]:
    """Parse a CSV blob into a list of dicts keyed by normalized headers."""
    text = blob.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []
    headers = [_normalize_header(h) for h in rows[0]]
    out: list[dict[str, Any]] = []
    for raw in rows[1:]:
        if not any((str(c).strip() if c is not None else "") for c in raw):
            continue  # skip blank lines
        d: dict[str, Any] = {}
        for i, key in enumerate(headers):
            d[key] = raw[i] if i < len(raw) else ""
        out.append(d)
    return out


def parse_xlsx_bytes(blob: bytes) -> list[dict[str, Any]]:
    """Parse an XLSX blob (first sheet) into dicts keyed by normalized headers."""
    wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    headers = [_normalize_header(h) for h in header_row]
    out: list[dict[str, Any]] = []
    for raw in rows_iter:
        if raw is None:
            continue
        if not any(("" if c is None else str(c).strip()) for c in raw):
            continue
        d: dict[str, Any] = {}
        for i, key in enumerate(headers):
            d[key] = raw[i] if i < len(raw) else ""
        out.append(d)
    return out


def parse_upload(filename: str, blob: bytes) -> list[dict[str, Any]]:
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return parse_xlsx_bytes(blob)
    # Default to CSV
    return parse_csv_bytes(blob)


def validate_row(
    row: dict[str, Any],
    valid_categories: set[str],
) -> dict[str, Any]:
    """Convert a raw row dict into a normalized result with errors.

    Returns: {"ok": bool, "errors": [str], "data": {...}, "mode": "create"|"update"}
    """
    errors: list[str] = []
    product_id = (str(row.get("product_id") or "")).strip() or None
    mode = "update" if product_id else "create"

    name = (str(row.get("name") or "")).strip()
    description = (str(row.get("description") or "")).strip()
    category = (str(row.get("category") or "")).strip()
    subcategory = (str(row.get("subcategory") or "")).strip() or None
    price_nzd = _coerce_float(row.get("price_nzd"))
    stock_count = _coerce_int(row.get("stock_count"), default=None)
    sizes = _split_multi(row.get("sizes"))[:12]
    colors = _split_multi(row.get("colors"))[:10]
    smin = _coerce_int(row.get("shipping_days_min"), default=7) or 7
    smax = _coerce_int(row.get("shipping_days_max"), default=14) or 14
    images = _split_multi(row.get("image_urls"))[:10]

    if mode == "create":
        if not name or len(name) < 2:
            errors.append("name is required (min 2 chars)")
        if not description or len(description) < 10:
            errors.append("description is required (min 10 chars)")
        if not category:
            errors.append("category is required")
        elif valid_categories and category not in valid_categories:
            errors.append(f"category '{category}' is not a known taxonomy category")
        if price_nzd is None or price_nzd <= 0:
            errors.append("price_nzd must be a positive number")
        if stock_count is None or stock_count < 0:
            errors.append("stock_count is required (>= 0)")
        if not images:
            errors.append("image_urls is required (1+ URLs, separated by | or ,)")
    else:
        # Update mode — only validate fields that are present.
        if name and len(name) < 2:
            errors.append("name must be at least 2 chars")
        if description and len(description) < 10:
            errors.append("description must be at least 10 chars")
        if category and valid_categories and category not in valid_categories:
            errors.append(f"category '{category}' is not a known taxonomy category")
        if price_nzd is not None and price_nzd <= 0:
            errors.append("price_nzd must be positive")
        if stock_count is not None and stock_count < 0:
            errors.append("stock_count must be >= 0")

    # Basic image URL sanity
    for u in images:
        if not (u.startswith("http://") or u.startswith("https://") or u.startswith("data:")):
            errors.append(f"image_url '{u[:40]}…' is not a valid URL")
            break

    if smin < 1 or smax < 1 or smin > smax:
        errors.append("shipping_days_min/max are invalid (min must be <= max, both >= 1)")

    data = {
        "product_id": product_id,
        "name": name,
        "description": description,
        "category": category,
        "subcategory": subcategory,
        "price_nzd": price_nzd,
        "stock_count": stock_count,
        "sizes": sizes,
        "colors": colors,
        "shipping_days_min": smin,
        "shipping_days_max": smax,
        "images": images,
    }
    return {"ok": not errors, "errors": errors, "data": data, "mode": mode}


# ---------------------------------------------------------------------------
# CSV/XLSX writers
# ---------------------------------------------------------------------------
def write_csv(rows: Iterable[dict[str, Any]], columns: list[str] | None = None) -> str:
    cols = columns or TEMPLATE_COLUMNS
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(cols)
    for r in rows:
        writer.writerow([_stringify_cell(r.get(c, "")) for c in cols])
    return buf.getvalue()


def write_xlsx(rows: Iterable[dict[str, Any]], columns: list[str] | None = None) -> bytes:
    cols = columns or TEMPLATE_COLUMNS
    wb = Workbook()
    ws = wb.active
    ws.title = "Listings"
    ws.append(cols)
    for r in rows:
        ws.append([_stringify_cell(r.get(c, "")) for c in cols])
    # Friendly default widths
    for i, c in enumerate(cols, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(14, min(40, len(c) + 6))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _stringify_cell(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, list):
        return " | ".join(str(x) for x in v)
    return v


def template_rows_example() -> list[dict[str, Any]]:
    """Two example rows shown inside the downloaded template."""
    return [
        {
            "product_id": "",
            "name": "Banarasi Silk Saree",
            "description": "Hand-woven Banarasi silk saree with zari work and matching blouse.",
            "category": "Ethnic Fashion",
            "subcategory": "Sarees",
            "price_nzd": 189.0,
            "stock_count": 12,
            "sizes": "Free Size",
            "colors": "Red | Maroon",
            "shipping_days_min": 7,
            "shipping_days_max": 14,
            "image_urls": "https://example.com/saree-front.jpg | https://example.com/saree-back.jpg",
        },
        {
            "product_id": "",
            "name": "Brass Diya Set (Pack of 4)",
            "description": "Traditional handcrafted brass diyas for pooja and festive decoration.",
            "category": "Home & Puja",
            "subcategory": "Idols",
            "price_nzd": 32.5,
            "stock_count": 50,
            "sizes": "",
            "colors": "Gold",
            "shipping_days_min": 5,
            "shipping_days_max": 10,
            "image_urls": "https://example.com/diya.jpg",
        },
    ]
