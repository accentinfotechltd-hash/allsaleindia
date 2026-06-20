"""Tests for bulk listing CSV/XLSX upload, preview & import."""
import io
import os
import time
from pathlib import Path

import pytest
import requests
from openpyxl import Workbook, load_workbook

BASE_URL = (os.environ.get("EXPO_PUBLIC_BACKEND_URL") or "").rstrip("/")
if not BASE_URL:
    for line in Path("/app/frontend/.env").read_text().splitlines():
        if line.startswith("EXPO_PUBLIC_BACKEND_URL="):
            BASE_URL = line.split("=", 1)[1].strip().strip('"').rstrip("/")
            break


def _gstin_pan():
    from _helpers import make_gstin_pan

    return make_gstin_pan()


@pytest.fixture(scope="module")
def seller_session():
    """Create a verified seller and return its headers."""
    g, p = _gstin_pan()
    suffix = int(time.time() * 1000)
    email = f"TEST_bulk_seller_{suffix}@allsale.co.nz"
    body = {
        "email": email,
        "password": "Test1234!",
        "business": {
            "business_type": "private_limited",
            "company_name": "TEST Bulk Imports Pvt Ltd",
            "gstin": g,
            "pan": p,
            "cin": "U74999MH2020PTC123456",
            "address_line1": "9 Bulk Lane",
            "address_line2": "Andheri",
            "city": "Mumbai",
            "state": "Maharashtra",
            "pincode": "400001",
            "contact_name": "Bulk Tester",
            "contact_phone": "+911234567890",
        },
    }
    r = requests.post(f"{BASE_URL}/api/seller/register", json=body)
    assert r.status_code == 200, r.text
    token = r.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    # Fast-forward verification — admin manually approves in real life;
    # for tests we flip the seller_verification_status straight to
    # "approved" on the user document and the seller's profile so the
    # full upload/import suite can run end-to-end.
    user_id = r.json()["user"]["id"]
    try:
        from pymongo import MongoClient
        cli = MongoClient(os.environ.get("MONGO_URL", "mongodb://localhost:27017"))
        db_sync = cli[os.environ.get("DB_NAME", "allsale_database")]
        db_sync.users.update_one(
            {"id": user_id},
            {"$set": {"seller_verification_status": "approved"}},
        )
        db_sync.sellers.update_one(
            {"user_id": user_id},
            {"$set": {"verification_status": "approved"}},
        )
    except Exception:
        pass

    return headers


def test_csv_template_download(seller_session):
    r = requests.get(
        f"{BASE_URL}/api/seller/bulk/template.csv", headers=seller_session
    )
    assert r.status_code == 200
    assert "text/csv" in r.headers["content-type"]
    body = r.text
    assert "name" in body and "price_nzd" in body
    # Has example rows
    assert "Banarasi" in body or "Saree" in body


def test_xlsx_template_download(seller_session):
    r = requests.get(
        f"{BASE_URL}/api/seller/bulk/template.xlsx", headers=seller_session
    )
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    assert "name" in headers
    assert "price_nzd" in headers
    assert len(rows) >= 2  # header + at least one example row


def test_columns_endpoint(seller_session):
    r = requests.get(
        f"{BASE_URL}/api/seller/bulk/columns", headers=seller_session
    )
    assert r.status_code == 200
    data = r.json()
    assert "name" in data["columns"]
    assert "price_nzd" in data["columns"]
    assert isinstance(data["categories"], list) and len(data["categories"]) > 0
    assert data["max_rows_per_upload"] >= 100


def _build_csv(rows: list[dict]) -> bytes:
    cols = list(rows[0].keys())
    out = ",".join(cols) + "\n"
    for r in rows:
        out += ",".join(f'"{(r[c] or "")}"' for c in cols) + "\n"
    return out.encode()


def test_preview_valid_csv(seller_session):
    csv_bytes = _build_csv(
        [
            {
                "product_id": "",
                "name": "Test Bulk Item A",
                "description": "A great test product for bulk import flow.",
                "category": "Ethnic Fashion",
                "subcategory": "Sarees",
                "price_nzd": "45.5",
                "stock_count": "20",
                "sizes": "S | M | L",
                "colors": "Red | Blue",
                "shipping_days_min": "7",
                "shipping_days_max": "14",
                "image_urls": "https://example.com/a.jpg",
            },
            {
                "product_id": "",
                "name": "Test Bulk Item B",
                "description": "Another great test product for bulk import.",
                "category": "Home & Puja",
                "subcategory": "Brass Items",
                "price_nzd": "120",
                "stock_count": "5",
                "sizes": "",
                "colors": "",
                "shipping_days_min": "5",
                "shipping_days_max": "10",
                "image_urls": "https://example.com/b1.jpg | https://example.com/b2.jpg",
            },
        ]
    )
    files = {"file": ("listings.csv", csv_bytes, "text/csv")}
    r = requests.post(
        f"{BASE_URL}/api/seller/bulk/preview",
        files=files,
        headers=seller_session,
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["total"] == 2
    assert data["valid"] == 2
    assert data["will_create"] == 2
    assert data["will_update"] == 0
    assert data["errors"] == 0
    assert all(r["ok"] for r in data["rows"])


def test_preview_validation_errors(seller_session):
    csv_bytes = _build_csv(
        [
            {
                "product_id": "",
                "name": "X",  # too short
                "description": "short",  # too short
                "category": "Made-up Category",  # not a real category
                "price_nzd": "-5",  # negative
                "stock_count": "",  # missing
                "image_urls": "not-a-url",  # bad URL
            }
        ]
    )
    files = {"file": ("bad.csv", csv_bytes, "text/csv")}
    r = requests.post(
        f"{BASE_URL}/api/seller/bulk/preview",
        files=files,
        headers=seller_session,
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["total"] == 1
    assert data["errors"] == 1
    assert data["valid"] == 0
    row = data["rows"][0]
    assert row["ok"] is False
    errs = " ".join(row["errors"]).lower()
    assert "name" in errs
    assert "price" in errs
    assert "image" in errs or "url" in errs


def test_full_import_creates_listings(seller_session):
    """End-to-end: preview a CSV then commit it, listings show up."""
    csv_bytes = _build_csv(
        [
            {
                "product_id": "",
                "name": "Bulk Test Saree One",
                "description": "Lovely test silk saree for end-to-end import.",
                "category": "Ethnic Fashion",
                "subcategory": "Sarees",
                "price_nzd": "99.99",
                "stock_count": "10",
                "sizes": "Free Size",
                "colors": "Red",
                "shipping_days_min": "7",
                "shipping_days_max": "14",
                "image_urls": "https://example.com/saree-one.jpg",
            },
            {
                "product_id": "",
                "name": "Bulk Test Saree Two",
                "description": "Another beautiful test silk saree for import.",
                "category": "Ethnic Fashion",
                "subcategory": "Sarees",
                "price_nzd": "149.50",
                "stock_count": "8",
                "sizes": "Free Size",
                "colors": "Blue",
                "shipping_days_min": "7",
                "shipping_days_max": "14",
                "image_urls": "https://example.com/saree-two.jpg",
            },
        ]
    )
    files = {"file": ("good.csv", csv_bytes, "text/csv")}
    pr = requests.post(
        f"{BASE_URL}/api/seller/bulk/preview",
        files=files,
        headers=seller_session,
    )
    assert pr.status_code == 200
    preview = pr.json()
    assert preview["valid"] == 2
    # Pull rows out (server already normalized them)
    rows = [r["data"] for r in preview["rows"] if r["ok"]]

    ir = requests.post(
        f"{BASE_URL}/api/seller/bulk/import",
        json={"rows": rows},
        headers={**seller_session, "Content-Type": "application/json"},
    )
    assert ir.status_code == 200, ir.text
    result = ir.json()
    assert result["created"] == 2
    assert result["updated"] == 0
    assert result["total_attempted"] == 2
    assert result["errors"] == []

    # Listings now visible via /seller/products
    lr = requests.get(f"{BASE_URL}/api/seller/products", headers=seller_session)
    assert lr.status_code == 200
    names = [p["name"] for p in lr.json()]
    assert "Bulk Test Saree One" in names
    assert "Bulk Test Saree Two" in names


def test_round_trip_export_edit_import(seller_session):
    """Export current listings as CSV, change a price, re-import → update."""
    # 1) Export
    er = requests.get(
        f"{BASE_URL}/api/seller/bulk/export.csv", headers=seller_session
    )
    assert er.status_code == 200
    body = er.text
    assert "product_id" in body.split("\n")[0]

    # Parse with stdlib so we can edit the price of "Bulk Test Saree One".
    import csv

    reader = csv.DictReader(io.StringIO(body))
    rows = list(reader)
    assert any(r["name"] == "Bulk Test Saree One" for r in rows)
    target = next(r for r in rows if r["name"] == "Bulk Test Saree One")
    assert target["product_id"]  # not empty
    target["price_nzd"] = "55.55"  # change the price
    target["stock_count"] = "3"

    # Re-write only that single row to upload (smaller payload).
    out = io.StringIO()
    writer = csv.DictWriter(out, fieldnames=reader.fieldnames)
    writer.writeheader()
    writer.writerow(target)

    files = {"file": ("edit.csv", out.getvalue().encode(), "text/csv")}
    pr = requests.post(
        f"{BASE_URL}/api/seller/bulk/preview",
        files=files,
        headers=seller_session,
    )
    assert pr.status_code == 200, pr.text
    preview = pr.json()
    assert preview["will_update"] == 1
    assert preview["will_create"] == 0
    assert preview["valid"] == 1

    edit_rows = [r["data"] for r in preview["rows"] if r["ok"]]
    ir = requests.post(
        f"{BASE_URL}/api/seller/bulk/import",
        json={"rows": edit_rows},
        headers={**seller_session, "Content-Type": "application/json"},
    )
    assert ir.status_code == 200, ir.text
    assert ir.json()["updated"] == 1
    assert ir.json()["created"] == 0

    # Confirm the price actually changed
    lr = requests.get(f"{BASE_URL}/api/seller/products", headers=seller_session)
    item = next(p for p in lr.json() if p["name"] == "Bulk Test Saree One")
    assert item["price_nzd"] == 55.55
    assert item["stock_count"] == 3


def test_import_rejects_other_sellers_product_id(seller_session):
    """Spoofing a product_id from another seller must be rejected."""
    fake_id = "not-our-product-12345"
    rows = [
        {
            "product_id": fake_id,
            "name": "Spoofed Update",
            "description": "Trying to update someone else's product.",
            "category": "Ethnic Fashion",
            "price_nzd": 1.0,
            "stock_count": 1,
            "sizes": [],
            "colors": [],
            "shipping_days_min": 7,
            "shipping_days_max": 14,
            "images": ["https://example.com/x.jpg"],
        }
    ]
    ir = requests.post(
        f"{BASE_URL}/api/seller/bulk/import",
        json={"rows": rows},
        headers={**seller_session, "Content-Type": "application/json"},
    )
    assert ir.status_code == 200, ir.text
    res = ir.json()
    assert res["created"] == 0
    assert res["updated"] == 0
    assert len(res["errors"]) == 1
    assert any("not one of your listings" in e for e in res["errors"][0]["errors"])


def test_xlsx_upload(seller_session):
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "product_id",
            "name",
            "description",
            "category",
            "price_nzd",
            "stock_count",
            "image_urls",
        ]
    )
    ws.append(
        [
            "",
            "XLSX Test Product",
            "Imported from XLSX file end-to-end test product.",
            "Jewelry & Accessories",
            29.95,
            7,
            "https://example.com/xlsx-product.jpg",
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)
    files = {
        "file": (
            "listings.xlsx",
            buf.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    pr = requests.post(
        f"{BASE_URL}/api/seller/bulk/preview",
        files=files,
        headers=seller_session,
    )
    assert pr.status_code == 200, pr.text
    preview = pr.json()
    assert preview["valid"] == 1
    assert preview["will_create"] == 1
    rows = [r["data"] for r in preview["rows"] if r["ok"]]
    ir = requests.post(
        f"{BASE_URL}/api/seller/bulk/import",
        json={"rows": rows},
        headers={**seller_session, "Content-Type": "application/json"},
    )
    assert ir.status_code == 200
    assert ir.json()["created"] == 1
