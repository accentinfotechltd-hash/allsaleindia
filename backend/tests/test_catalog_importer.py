"""Tests for the catalog importer service (Amazon, Flipkart, CSV).

These tests build small synthetic xlsx/xls/csv fixtures in-memory so they
don't depend on third-party files. The real Amazon/Flipkart files were
used during development to validate the row-detection heuristics.
"""
from __future__ import annotations

import io

import openpyxl
import pytest

from services.catalog_importer import detect_format, parse
from services.catalog_importer.mapping import (
    coerce_inr_to_nzd,
    map_flipkart_sheet,
    map_amazon_product_type,
    parse_decimal,
    split_multi,
)


# ---------------------------------------------------------------------------
# mapping.py
# ---------------------------------------------------------------------------
def test_parse_decimal_handles_currency_and_commas():
    assert parse_decimal("₹1,234.50") == 1234.5
    assert parse_decimal("$999") == 999.0
    assert parse_decimal("INR 100") == 100.0
    assert parse_decimal(None) is None
    assert parse_decimal("") is None
    assert parse_decimal("abc") is None
    assert parse_decimal(42) == 42.0


def test_split_multi_supports_flipkart_amazon_separators():
    assert split_multi("a::b::c") == ["a", "b", "c"]
    assert split_multi("a|b|c") == ["a", "b", "c"]
    assert split_multi("a;b;c") == ["a", "b", "c"]
    assert split_multi("solo") == ["solo"]
    assert split_multi("") == []
    assert split_multi(None) == []


def test_coerce_inr_to_nzd_default_fx_when_zero():
    assert coerce_inr_to_nzd(5100, 0) == 102.0  # falls back to /50
    assert coerce_inr_to_nzd(5100, 51) == 100.0


def test_map_flipkart_sheet_known_and_fuzzy():
    assert map_flipkart_sheet("conditioner") == ("Beauty & Health", "Hair Care")
    assert map_flipkart_sheet("Saree") == ("Ethnic Fashion", "Sarees")
    assert map_flipkart_sheet("super-saree-pack") == ("Ethnic Fashion", "Sarees")
    assert map_flipkart_sheet("unknown_widget") == (None, None)


def test_map_amazon_product_type_partial_matches():
    cat, _ = map_amazon_product_type("HAIR_CONDITIONER")
    assert cat == "Beauty & Health"
    cat, _ = map_amazon_product_type("WIRELESS_ACCESSORY")
    assert cat == "Electronics"
    cat, _ = map_amazon_product_type("RANDOM_FOOD_KIT")
    assert cat == "Food & Groceries"
    cat, _ = map_amazon_product_type("XYZ123")
    assert cat is None


# ---------------------------------------------------------------------------
# detect.py
# ---------------------------------------------------------------------------
def test_detect_csv_by_extension():
    assert detect_format("catalog.csv", b"sku,name\n1,foo") == "csv"
    assert detect_format("catalog.tsv", b"sku\tname\n1\tfoo") == "csv"


def test_detect_unknown_for_short_garbage():
    assert detect_format("a.bin", b"\x00\x00") == "unknown"


# ---------------------------------------------------------------------------
# Amazon parser
# ---------------------------------------------------------------------------
def _build_amazon_sheet() -> bytes:
    """Build a minimal Amazon-template-shaped workbook with 1 row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    # Rows 1-3: control headers (we skip them)
    ws.cell(row=1, column=1, value="settings=feedType=...")
    ws.cell(row=3, column=4, value="Variations")
    # Row 4: human-readable labels (the parser keys on these)
    headers = [
        "SKU", "Product Type", "Item Name", "Brand Name",
        "Main Image URL", "Other Image URL", "Other Image URL",
        "Product Description", "Bullet Point", "Bullet Point",
        "Color", "Size", "Manufacturer", "Standard Price",
        "Quantity", "Country of Origin",
    ]
    for i, h in enumerate(headers):
        ws.cell(row=4, column=i + 1, value=h)
    # Row 5: internal field codes (parser skips)
    for i in range(len(headers)):
        ws.cell(row=5, column=i + 1, value="code_placeholder")
    # Row 6: real product
    ws.cell(row=6, column=1, value="SKU-A")
    ws.cell(row=6, column=2, value="HAIR_CONDITIONER")
    ws.cell(row=6, column=3, value="Silky Smooth Conditioner 250ml")
    ws.cell(row=6, column=4, value="Allsale Beauty")
    ws.cell(row=6, column=5, value="https://cdn.example.com/main.jpg")
    ws.cell(row=6, column=6, value="https://cdn.example.com/g1.jpg")
    ws.cell(row=6, column=7, value="https://cdn.example.com/g2.jpg")
    ws.cell(row=6, column=8, value="A nourishing conditioner with coconut and almond.")
    ws.cell(row=6, column=9, value="Coconut oil")
    ws.cell(row=6, column=10, value="Tames frizz")
    ws.cell(row=6, column=11, value="Black, White")
    ws.cell(row=6, column=12, value="250ml, 500ml")
    ws.cell(row=6, column=13, value="Allsale Beauty Pvt Ltd")
    ws.cell(row=6, column=14, value=299.00)
    ws.cell(row=6, column=15, value=50)
    ws.cell(row=6, column=16, value="India")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def test_amazon_parser_extracts_full_product_row():
    blob = _build_amazon_sheet()
    fmt = detect_format("amazon-template.xlsx", blob)
    # xlsx without 'amazon' in name defaults to 'amazon' only for .xlsm
    # Force the parser path explicitly:
    res = parse(blob, "amazon")
    assert res.source == "amazon"
    assert res.total_rows == 1
    assert res.ready_count == 1
    row = res.rows[0]
    p = row.product
    assert p.sku == "SKU-A"
    assert p.name.startswith("Silky Smooth")
    assert p.brand == "Allsale Beauty"
    assert p.category == "Beauty & Health"
    assert p.image and p.image.startswith("https://cdn.")
    assert len(p.images) == 2
    assert len(p.bullets) == 2
    assert p.colors == ["Black", "White"]
    assert p.sizes == ["250ml", "500ml"]
    assert p.price_inr == 299.0
    assert p.price_nzd is not None and 5.0 < p.price_nzd < 7.0
    assert p.stock_count == 50


# ---------------------------------------------------------------------------
# Flipkart parser
# ---------------------------------------------------------------------------
def _build_flipkart_sheet() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "saree"
    headers = [
        "Flipkart Serial Number", "Catalog QC Status", "_a", "_b", "_c",
        "_d", "Seller SKU ID", "Group ID", "Parent Variant FSN", "_e",
        "Listing Status", "MRP (INR)", "Your selling price (INR)",
        "Fullfilment by", "Procurement type", "Procurement SLA (DAY)",
        "Stock", "Shipping provider", "_f", "_g", "_h",
        "Length (CM)", "Breadth (CM)", "Height (CM)", "Weight (KG)",
        "HSN", "_lc", "Country Of Origin", "Manufacturer Details", "_pd",
        "Importer Details", "_md", "_sl", "_tc", "_minoq", "Brand",
        "Model Name", "_idf", "_st", "_pt",
        "Main Image URL", "Other Image URL 1", "Other Image URL 2",
        "Other Image URL 3", "Other Image URL 4",
        "_q1", "_q2", "EAN/UPC", "Description", "Key Features", "_kw",
    ]
    for i, h in enumerate(headers):
        ws.cell(row=1, column=i + 1, value=h)
    # Rows 2,3,4 = meta (parser skips)
    for r in (2, 3, 4):
        ws.cell(row=r, column=1, value="meta")

    def cidx(label):
        return headers.index(label) + 1

    # Row 5: good product
    ws.cell(row=5, column=cidx("Seller SKU ID"), value="BNRS-001")
    ws.cell(row=5, column=cidx("MRP (INR)"), value=2999)
    ws.cell(row=5, column=cidx("Your selling price (INR)"), value=1499)
    ws.cell(row=5, column=cidx("Stock"), value=42)
    ws.cell(row=5, column=cidx("Length (CM)"), value=30)
    ws.cell(row=5, column=cidx("Breadth (CM)"), value=20)
    ws.cell(row=5, column=cidx("Height (CM)"), value=5)
    ws.cell(row=5, column=cidx("Weight (KG)"), value=0.4)
    ws.cell(row=5, column=cidx("HSN"), value="50079010")
    ws.cell(row=5, column=cidx("Country Of Origin"), value="India")
    ws.cell(row=5, column=cidx("Brand"), value="Banaras Weaves")
    ws.cell(row=5, column=cidx("Model Name"), value="Tangail Silk")
    ws.cell(row=5, column=cidx("Main Image URL"), value="https://cdn.test/main.jpg")
    ws.cell(row=5, column=cidx("Other Image URL 1"), value="https://cdn.test/1.jpg")
    ws.cell(row=5, column=cidx("EAN/UPC"), value="8901234567893")
    ws.cell(row=5, column=cidx("Description"), value="Hand-woven Banarasi silk saree.")
    ws.cell(row=5, column=cidx("Key Features"), value="Pure silk::Hand woven::Zari work")

    # Row 6: missing image -> should error
    ws.cell(row=6, column=cidx("Brand"), value="Bad")
    ws.cell(row=6, column=cidx("Model Name"), value="No-Image")
    ws.cell(row=6, column=cidx("MRP (INR)"), value=100)
    ws.cell(row=6, column=cidx("Your selling price (INR)"), value=50)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def test_flipkart_parser_with_seller_rows():
    blob = _build_flipkart_sheet()
    res = parse(blob, "flipkart")
    assert res.source == "flipkart"
    assert res.sheet_name == "saree"
    assert res.total_rows == 2
    assert res.ready_count == 1
    assert res.needs_attention_count == 1

    good = res.rows[0]
    p = good.product
    assert good.ready is True
    assert p.name.startswith("Banaras Weaves Tangail")
    assert p.category == "Ethnic Fashion"
    assert p.subcategory == "Sarees"
    assert p.price_inr == 1499.0
    assert p.mrp_inr == 2999.0
    assert p.stock_count == 42
    assert p.hsn_code == "50079010"
    assert p.ean_upc == "8901234567893"
    assert p.weight_kg == 0.4
    assert p.dimensions_cm == [30.0, 20.0, 5.0]
    assert p.bullets == ["Pure silk", "Hand woven", "Zari work"]

    bad = res.rows[1]
    assert bad.ready is False
    assert any(i.severity == "error" and i.field == "image" for i in bad.issues)


# ---------------------------------------------------------------------------
# CSV parser
# ---------------------------------------------------------------------------
def test_csv_parser_accepts_basic_template():
    csv_body = (
        b"sku,name,description,category,price_nzd,stock_count,image,bullets\n"
        b"X1,Cool Saree,Hand woven,Ethnic Fashion,29.99,10,https://x/y.jpg,Soft;Light;Quick-dry\n"
        b"X2,,,Ethnic Fashion,9.99,5,,\n"
    )
    res = parse(csv_body, "csv")
    assert res.total_rows == 2
    assert res.ready_count == 1
    assert res.needs_attention_count == 1
    good = res.rows[0].product
    assert good.sku == "X1"
    assert good.price_nzd == 29.99
    assert good.bullets == ["Soft", "Light", "Quick-dry"]
    bad = res.rows[1]
    assert not bad.ready
    assert any(i.severity == "error" for i in bad.issues)


# ---------------------------------------------------------------------------
# Unknown / unsupported
# ---------------------------------------------------------------------------
def test_parse_raises_on_unknown_format():
    with pytest.raises(ValueError):
        parse(b"junk", "weirdmarketplace")
